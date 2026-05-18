"""
One-time database schema initialisation for the AURA ingestion pipeline.

Run this script ONCE before executing the ingestion job for the first time,
or whenever the schema needs to be re-created on a fresh database.

Usage:
    cd jobs/sharepoint_ingestion
    python create_schema.py
"""
import sys
import os
import re
import datetime

_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

import psycopg2
import openpyxl
from config.settings import settings
from utils.logging_config import get_logger

_EXCEL_PATH = os.path.join(_HERE, "data", "zoho_dummy_data.xlsx")

logger = get_logger("create_schema")

_SCHEMA_SQL = """
-- =========================================================
-- AURA ENTERPRISE AI ASSISTANT PLATFORM
-- PostgreSQL + pgvector Schema
-- =========================================================

CREATE EXTENSION IF NOT EXISTS pgcrypto;
CREATE EXTENSION IF NOT EXISTS vector;


-- =========================================================
-- TABLE: users
-- =========================================================
CREATE TABLE IF NOT EXISTS users (
    id          UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    email       VARCHAR(255) UNIQUE NOT NULL,
    azure_oid   VARCHAR(255) UNIQUE,
    display_name VARCHAR(255),
    role        VARCHAR(50) NOT NULL DEFAULT 'user',
                -- user | admin | auditor
    is_active   BOOLEAN NOT NULL DEFAULT TRUE,
    last_login_at TIMESTAMP,
    created_at  TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at  TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_users_email  ON users(email);
CREATE INDEX IF NOT EXISTS idx_users_active ON users(is_active);


-- =========================================================
-- TABLE: user_sessions
-- =========================================================
CREATE TABLE IF NOT EXISTS user_sessions (
    id            UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    user_id       UUID NOT NULL REFERENCES users(id) ON DELETE CASCADE,
    session_id    VARCHAR(255) UNIQUE NOT NULL,
    ip_address    VARCHAR(45),
    user_agent    TEXT,
    login_time    TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    last_activity TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    is_active     BOOLEAN NOT NULL DEFAULT TRUE
);

CREATE INDEX IF NOT EXISTS idx_user_sessions_user   ON user_sessions(user_id);
CREATE INDEX IF NOT EXISTS idx_user_sessions_active ON user_sessions(is_active);


-- =========================================================
-- TABLE: conversations
-- =========================================================
CREATE TABLE IF NOT EXISTS conversations (
    id          UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    user_id     UUID NOT NULL REFERENCES users(id) ON DELETE CASCADE,
    title       VARCHAR(500),
    is_pinned   BOOLEAN NOT NULL DEFAULT FALSE,
    is_archived BOOLEAN NOT NULL DEFAULT FALSE,
    is_deleted  BOOLEAN NOT NULL DEFAULT FALSE,
                -- soft delete for compliance
    created_at  TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at  TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_conversations_user_updated
    ON conversations(user_id, updated_at DESC)
    WHERE is_deleted = FALSE;


-- =========================================================
-- TABLE: messages
-- =========================================================
CREATE TABLE IF NOT EXISTS messages (
    id                UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    conversation_id   UUID NOT NULL REFERENCES conversations(id) ON DELETE CASCADE,
    parent_message_id UUID REFERENCES messages(id),
                -- for regenerations / branching
    role              VARCHAR(20) NOT NULL,
                -- user | assistant | system | tool
    content           TEXT NOT NULL,
    agent_name        VARCHAR(100),
    model_name        VARCHAR(100),
                -- which underlying LLM answered
    route_type        VARCHAR(50),
                -- rag | api | mixed
    confidence_score  NUMERIC(5,2),
    token_usage       INTEGER,
    response_time_ms  INTEGER,
    citations         JSONB,
    status            VARCHAR(30) NOT NULL DEFAULT 'complete',
                -- streaming | complete | failed | cancelled
    created_at        TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,

    CONSTRAINT chk_messages_role
        CHECK (role IN ('user','assistant','system','tool'))
);

CREATE INDEX IF NOT EXISTS idx_messages_conversation
    ON messages(conversation_id, created_at);
CREATE INDEX IF NOT EXISTS idx_messages_parent
    ON messages(parent_message_id);


-- =========================================================
-- TABLE: feedback
-- =========================================================
CREATE TABLE IF NOT EXISTS feedback (
    id          UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    message_id  UUID NOT NULL REFERENCES messages(id) ON DELETE CASCADE,
    user_id     UUID NOT NULL REFERENCES users(id),
    rating      VARCHAR(10) NOT NULL,
                -- up | down
    category    VARCHAR(50),
                -- incorrect | incomplete | hallucination | harmful | formatting | other
    comment     TEXT,
    created_at  TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,

    CONSTRAINT chk_feedback_rating CHECK (rating IN ('up','down')),
    CONSTRAINT uq_feedback_per_message UNIQUE (message_id, user_id)
);

CREATE INDEX IF NOT EXISTS idx_feedback_message ON feedback(message_id);
CREATE INDEX IF NOT EXISTS idx_feedback_rating  ON feedback(rating, created_at DESC);


-- =========================================================
-- TABLE: documents
-- =========================================================
CREATE TABLE IF NOT EXISTS documents (
    id              UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    source_system   VARCHAR(100) NOT NULL,
                -- sharepoint | nexus
    document_name   TEXT NOT NULL,
    source_path     TEXT UNIQUE NOT NULL,
    document_type   VARCHAR(50),
                -- pdf | pptx | docx | xlsx
    checksum        VARCHAR(255),
    visibility      VARCHAR(50) NOT NULL DEFAULT 'all',
                -- all | hr | finance | it | restricted
    tags            JSONB,
    last_modified   TIMESTAMP,
    indexed_at      TIMESTAMP,
    status          VARCHAR(50) NOT NULL DEFAULT 'pending',
                -- pending | indexed | failed
    created_at      TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at      TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_documents_source_path ON documents(source_path);
CREATE INDEX IF NOT EXISTS idx_documents_status      ON documents(status);
CREATE INDEX IF NOT EXISTS idx_documents_visibility  ON documents(visibility);


-- =========================================================
-- TABLE: document_chunks
-- =========================================================
CREATE TABLE IF NOT EXISTS document_chunks (
    id               UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    document_id      UUID NOT NULL REFERENCES documents(id) ON DELETE CASCADE,
    chunk_index      INTEGER NOT NULL,
    chunk_hash       VARCHAR(255),
    chunk_text       TEXT NOT NULL,
    embedding        VECTOR({dim}),
    embedding_model  VARCHAR(100),
                -- e.g. text-embedding-3-small
    page_number      INTEGER,
    section_heading  TEXT,
    metadata         JSONB,
    created_at       TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,

    CONSTRAINT uq_chunk_per_doc UNIQUE (document_id, chunk_index)
);

CREATE INDEX IF NOT EXISTS idx_document_chunks_doc
    ON document_chunks(document_id);

-- Vector index for cosine similarity search
CREATE INDEX IF NOT EXISTS idx_document_chunks_vector
    ON document_chunks
    USING ivfflat (embedding vector_cosine_ops)
    WITH (lists = 100);


-- =========================================================
-- TABLE: escalation_records
-- =========================================================
CREATE TABLE IF NOT EXISTS escalation_records (
    id              UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    user_id         UUID NOT NULL REFERENCES users(id),
    conversation_id UUID REFERENCES conversations(id),
    message_id      UUID REFERENCES messages(id),
    escalation_type VARCHAR(100) NOT NULL,
                -- hr | admin | it
    subject         VARCHAR(500) NOT NULL,
    reason          TEXT,
    form_payload    JSONB,
    priority        VARCHAR(20) NOT NULL DEFAULT 'medium',
                -- low | medium | high | critical
    status          VARCHAR(50) NOT NULL DEFAULT 'submitted',
                -- submitted | in_progress | resolved | closed
    assigned_team   VARCHAR(100),
    assigned_to     UUID REFERENCES users(id),
    resolved_at     TIMESTAMP,
    resolution_notes TEXT,
    created_at      TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at      TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_escalation_user      ON escalation_records(user_id);
CREATE INDEX IF NOT EXISTS idx_escalation_status    ON escalation_records(status);
CREATE INDEX IF NOT EXISTS idx_escalation_team      ON escalation_records(assigned_team);
CREATE INDEX IF NOT EXISTS idx_escalation_assignee  ON escalation_records(assigned_to);


-- =========================================================
-- TABLE: agent_routing_logs
-- =========================================================
CREATE TABLE IF NOT EXISTS agent_routing_logs (
    id                UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    message_id        UUID NOT NULL REFERENCES messages(id) ON DELETE CASCADE,
    detected_intent   VARCHAR(255),
    selected_agent    VARCHAR(100),
    route_type        VARCHAR(50),
                -- rag | api | mixed
    confidence_score  NUMERIC(5,2),
    alternative_agents JSONB,
                -- runner-up scores for A/B analysis
    routing_latency_ms INTEGER,
    created_at        TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_routing_message ON agent_routing_logs(message_id);
CREATE INDEX IF NOT EXISTS idx_routing_agent   ON agent_routing_logs(selected_agent);


-- =========================================================
-- TABLE: prompt_logs
-- =========================================================
CREATE TABLE IF NOT EXISTS prompt_logs (
    id                UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    message_id        UUID NOT NULL REFERENCES messages(id) ON DELETE CASCADE,
    agent_name        VARCHAR(100),
    model_name        VARCHAR(100),
    prompt_template   TEXT,
    retrieved_context JSONB,
    final_prompt      TEXT,
    model_response    TEXT,
    latency_ms        INTEGER,
    created_at        TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_prompt_logs_message ON prompt_logs(message_id);
CREATE INDEX IF NOT EXISTS idx_prompt_logs_agent   ON prompt_logs(agent_name, created_at DESC);


-- =========================================================
-- TABLE: error_logs
-- =========================================================
CREATE TABLE IF NOT EXISTS error_logs (
    id            UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    module_name   VARCHAR(255),
    error_type    VARCHAR(255),
    error_message TEXT,
    stack_trace   TEXT,
    payload       JSONB,
    user_id       UUID REFERENCES users(id),
    conversation_id UUID REFERENCES conversations(id),
    created_at    TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_error_logs_module ON error_logs(module_name, created_at DESC);
CREATE INDEX IF NOT EXISTS idx_error_logs_type   ON error_logs(error_type);


-- =========================================================
-- TABLE: audit_logs
-- =========================================================
CREATE TABLE IF NOT EXISTS audit_logs (
    id          UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    user_id     UUID REFERENCES users(id),
    action      VARCHAR(255) NOT NULL,
    entity_type VARCHAR(100),
    entity_id   VARCHAR(255),
    ip_address  VARCHAR(45),
    user_agent  TEXT,
    status      VARCHAR(20) NOT NULL DEFAULT 'success',
                -- success | failure
    metadata    JSONB,
    created_at  TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_audit_user   ON audit_logs(user_id, created_at DESC);
CREATE INDEX IF NOT EXISTS idx_audit_action ON audit_logs(action, created_at DESC);
CREATE INDEX IF NOT EXISTS idx_audit_entity ON audit_logs(entity_type, entity_id);


-- =========================================================
-- TABLE: pii_redaction_rules
-- =========================================================
CREATE TABLE IF NOT EXISTS pii_redaction_rules (
    id                UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    rule_name         VARCHAR(100) NOT NULL,
    rule_version      INTEGER NOT NULL DEFAULT 1,
    pii_type          VARCHAR(50) NOT NULL,
                -- EMAIL | PHONE | PAN | AADHAAR | SSN | CREDIT_CARD
                -- PERSON_NAME | IP_ADDRESS | EMPLOYEE_ID | DOB | ADDRESS
    detection_method  VARCHAR(30) NOT NULL,
                -- REGEX | NER | LLM | DICTIONARY | CUSTOM
    pattern           TEXT,
    replacement_token VARCHAR(100) NOT NULL DEFAULT '[REDACTED]',
    severity          VARCHAR(20) NOT NULL DEFAULT 'MEDIUM',
                -- LOW | MEDIUM | HIGH | CRITICAL
    is_active         BOOLEAN NOT NULL DEFAULT TRUE,
    description       TEXT,
    created_by        UUID REFERENCES users(id),
    created_at        TIMESTAMP NOT NULL DEFAULT NOW(),
    updated_at        TIMESTAMP NOT NULL DEFAULT NOW(),

    CONSTRAINT uq_rule_name_version UNIQUE (rule_name, rule_version)
);

CREATE INDEX IF NOT EXISTS idx_pii_rules_active ON pii_redaction_rules(is_active);
CREATE INDEX IF NOT EXISTS idx_pii_rules_type   ON pii_redaction_rules(pii_type);


-- =========================================================
-- TABLE: pii_redaction_logs
-- =========================================================
CREATE TABLE IF NOT EXISTS pii_redaction_logs (
    id                UUID PRIMARY KEY DEFAULT gen_random_uuid(),

    -- What was scanned
    source_type       VARCHAR(30) NOT NULL,
                -- USER_PROMPT | MODEL_RESPONSE | RETRIEVED_CONTEXT
                -- DOCUMENT_CHUNK | ESCALATION_FORM | FEEDBACK_COMMENT
    source_table      VARCHAR(50),
    source_id         UUID,

    -- Conversation/user context
    user_id           UUID REFERENCES users(id),
    conversation_id   UUID REFERENCES conversations(id),
    message_id        UUID REFERENCES messages(id),

    -- What was caught
    rule_id           UUID NOT NULL REFERENCES pii_redaction_rules(id),
    pii_type          VARCHAR(50) NOT NULL,
    detection_method  VARCHAR(30) NOT NULL,
    match_count       INTEGER NOT NULL DEFAULT 1,

    -- Safe metadata only (NEVER the actual PII value)
    value_hash        VARCHAR(64),
    value_length      INTEGER,
    match_positions   JSONB,
    confidence_score  NUMERIC(5,2),

    -- What was done
    action_taken      VARCHAR(30) NOT NULL,
                -- REDACTED | MASKED | TOKENIZED | BLOCKED
                -- FLAGGED_ONLY | ALLOWED_BY_POLICY
    replacement_token VARCHAR(100),

    -- Review workflow
    is_false_positive BOOLEAN,
    reviewed_by       UUID REFERENCES users(id),
    reviewed_at       TIMESTAMP,
    review_notes      TEXT,

    created_at        TIMESTAMP NOT NULL DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_pii_logs_user
    ON pii_redaction_logs(user_id, created_at DESC);
CREATE INDEX IF NOT EXISTS idx_pii_logs_conversation
    ON pii_redaction_logs(conversation_id);
CREATE INDEX IF NOT EXISTS idx_pii_logs_message
    ON pii_redaction_logs(message_id);
CREATE INDEX IF NOT EXISTS idx_pii_logs_source
    ON pii_redaction_logs(source_table, source_id);
CREATE INDEX IF NOT EXISTS idx_pii_logs_type_time
    ON pii_redaction_logs(pii_type, created_at DESC);
CREATE INDEX IF NOT EXISTS idx_pii_logs_rule
    ON pii_redaction_logs(rule_id);
CREATE INDEX IF NOT EXISTS idx_pii_logs_false_positive
    ON pii_redaction_logs(is_false_positive)
    WHERE is_false_positive IS NOT NULL;
CREATE INDEX IF NOT EXISTS idx_pii_logs_action
    ON pii_redaction_logs(action_taken);


-- =========================================================
-- TABLE: allocation_role_map
-- Maps employee designation → allocation board role.
-- Seeded by seed_allocation_roles.py; add new designations here as needed.
-- =========================================================
CREATE TABLE IF NOT EXISTS allocation_role_map (
    id          BIGSERIAL PRIMARY KEY,
    designation VARCHAR(255) NOT NULL UNIQUE,
    role        VARCHAR(50)  NOT NULL
                CHECK (role IN ('executive','business_lead','functional_lead','team_lead','employee','admin')),
    created_at  TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at  TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_alloc_role_desig ON allocation_role_map(designation);
""".format(dim=settings.EMBEDDING_DIMENSION)


def _to_snake_case(name: str) -> str:
    if not name or not str(name).strip():
        return ""
    name = str(name).strip()
    name = name.replace('%', '_pct')
    name = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    name = re.sub(r'_+', '_', name)
    return name.strip('_').lower()


def _infer_pg_type(values: list) -> str:
    """
    Returns a PG type only when ALL non-null, non-empty sampled values agree
    on the same Python type. Falls back to TEXT for mixed columns.
    """
    detected = None
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            t = "BOOLEAN"
        elif isinstance(v, datetime.datetime):
            t = "TIMESTAMP"
        elif isinstance(v, int):
            t = "BIGINT"
        elif isinstance(v, float):
            t = "DOUBLE PRECISION"
        else:
            return "TEXT"
        if detected is None:
            detected = t
        elif detected != t:
            return "TEXT"
    return detected or "TEXT"


def _read_sheet(wb, sheet_name: str):
    """Return (col_map, data_rows) where col_map is list of (snake_name, col_index)."""
    ws = wb[sheet_name]
    rows = list(ws.rows)
    if not rows:
        return [], []
    raw_headers = [cell.value for cell in rows[0]]
    col_map = []
    seen = set()
    for i, h in enumerate(raw_headers):
        snake = _to_snake_case(h) if h is not None else ""
        if not snake:
            continue
        # Deduplicate: append _2, _3 ... if name already used
        unique = snake
        suffix = 2
        while unique in seen:
            unique = f"{snake}_{suffix}"
            suffix += 1
        seen.add(unique)
        col_map.append((unique, i))
    data_rows = rows[1:]
    return col_map, data_rows


def _create_zoho_tables(conn):
    """Create Zoho tables and add any missing columns (idempotent)."""
    wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            table_name = _to_snake_case(sheet_name)
            col_map, data_rows = _read_sheet(wb, sheet_name)
            if not col_map:
                continue

            # Sample up to 50 rows for type inference
            col_values = {snake: [] for snake, _ in col_map}
            for row in data_rows[:50]:
                cells = [cell.value for cell in row]
                for snake, idx in col_map:
                    if idx < len(cells):
                        col_values[snake].append(cells[idx])

            col_types = {snake: _infer_pg_type(col_values[snake]) for snake, _ in col_map}

            col_names_set = {snake for snake, _ in col_map}
            pk_name = "_row_id" if "id" in col_names_set else "id"

            col_defs = ",\n    ".join(
                f"{snake} {col_types[snake]}" for snake, _ in col_map
            )
            create_sql = (
                f"CREATE TABLE IF NOT EXISTS {table_name} (\n"
                f"    {pk_name} BIGSERIAL PRIMARY KEY,\n"
                f"    {col_defs}\n"
                f");"
            )
            with conn.cursor() as cur:
                cur.execute(create_sql)
                logger.info(f"Table '{table_name}' created or already exists")

                # Add columns that are missing (schema evolution)
                for snake, _ in col_map:
                    cur.execute(
                        """
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = %s AND column_name = %s
                        """,
                        (table_name, snake),
                    )
                    if not cur.fetchone():
                        cur.execute(
                            f"ALTER TABLE {table_name} "
                            f"ADD COLUMN IF NOT EXISTS {snake} {col_types[snake]}"
                        )
                        logger.info(f"Added column '{snake}' to '{table_name}'")
    finally:
        wb.close()


def _coerce_value(v, pg_type: str):
    """Convert empty strings to None for non-TEXT columns to avoid type errors."""
    if pg_type == "TEXT":
        return v
    if v == "":
        return None
    return v


def _import_zoho_data(conn):
    """Insert Excel rows into Zoho tables; skips tables that already have data."""
    wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            table_name = _to_snake_case(sheet_name)
            col_map, data_rows = _read_sheet(wb, sheet_name)
            if not col_map:
                continue

            # Infer types from first 50 rows so we know which columns need coercion
            col_values = {snake: [] for snake, _ in col_map}
            for row in data_rows[:50]:
                cells = [cell.value for cell in row]
                for snake, idx in col_map:
                    if idx < len(cells):
                        col_values[snake].append(cells[idx])
            col_types = {snake: _infer_pg_type(col_values[snake]) for snake, _ in col_map}

            with conn.cursor() as cur:
                cur.execute(f"SELECT COUNT(*) FROM {table_name}")
                if cur.fetchone()[0] > 0:
                    logger.info(f"Table '{table_name}' already has data — skipping import")
                    continue

                col_names = [snake for snake, _ in col_map]
                placeholders = ", ".join(["%s"] * len(col_names))
                insert_sql = (
                    f"INSERT INTO {table_name} ({', '.join(col_names)}) "
                    f"VALUES ({placeholders})"
                )

                inserted = 0
                for row in data_rows:
                    cells = [cell.value for cell in row]
                    values = [
                        _coerce_value(
                            cells[idx] if idx < len(cells) else None,
                            col_types[snake],
                        )
                        for snake, idx in col_map
                    ]
                    if all(v is None for v in values):
                        continue
                    cur.execute(insert_sql, values)
                    inserted += 1

                logger.info(f"Imported {inserted} rows into '{table_name}'")
    finally:
        wb.close()


def create_schema():
    # ── Step 1: Create the 'aura' database if it doesn't exist ──────────────
    # CREATE DATABASE cannot run inside a transaction, so we connect to the
    # 'postgres' maintenance DB first, create 'aura', then reconnect.
    admin_url = re.sub(r'/[^/?#]+(\?.*)?$', r'/postgres\1', settings.database_url)

    logger.info("Ensuring database 'aura' exists...")
    admin_conn = psycopg2.connect(admin_url)
    admin_conn.autocommit = True
    try:
        with admin_conn.cursor() as cur:
            cur.execute("SELECT 1 FROM pg_database WHERE datname = 'aura'")
            if not cur.fetchone():
                cur.execute("CREATE DATABASE aura")
                logger.info("Database 'aura' created")
            else:
                logger.info("Database 'aura' already exists")
    finally:
        admin_conn.close()

    # ── Step 2: Connect to 'aura' and apply extensions + full schema ─────────
    aura_url = re.sub(r'/[^/?#]+(\?.*)?$', r'/aura\1', settings.database_url)
    logger.info(
        f"Connecting to {settings.SQL_HOST}:{settings.SQL_PORT}/aura "
        f"as {settings.SQL_USERNAME}"
    )
    conn = psycopg2.connect(aura_url)
    conn.autocommit = True
    try:
        with conn.cursor() as cur:
            cur.execute(_SCHEMA_SQL)
        logger.info("Schema created / verified successfully")
    except Exception as exc:
        logger.error(f"Schema creation failed: {exc}")
        raise
    finally:
        conn.close()

    # ── Step 3: Create Zoho tables and import data (idempotent) ──────────────
    logger.info("Setting up Zoho data tables...")
    conn = psycopg2.connect(aura_url)
    conn.autocommit = True
    try:
        _create_zoho_tables(conn)
        _import_zoho_data(conn)
        logger.info("Zoho tables ready")
    except Exception as exc:
        logger.error(f"Zoho table setup failed: {exc}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    create_schema()
